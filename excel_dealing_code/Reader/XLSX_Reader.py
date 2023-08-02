from Reader.Reader import Reader
import openpyxl


class XLSX_Reader(Reader):
    def __init__(self, file_name):
        super().__init__(file_name)
        self.wb = openpyxl.load_workbook(file_name, read_only=True)
        self.sht = self.wb.worksheets[0]

    def get_titles(self):
        head = self.sht[1]
        return [item.value for item in head]

    def close(self):
        self.wb.close()

    def get_row_number(self):
        return self.sht.max_row

    def travel_row(self, with_title=False):
        begin_row = 1 if with_title else 2
        for row_data in self.sht.iter_rows(min_row=begin_row):
            yield [item.value for item in row_data]

