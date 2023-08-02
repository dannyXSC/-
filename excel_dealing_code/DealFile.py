import openpyxl
from openpyxl import load_workbook
import xlrd

from DealIPC import deal_ipc
from tools import MY_DEBUG, LOG

need_col = ["公开（公告）号",
            "公开（公告）日",
            "申请日",
            "专利类型",
            "公开国别",
            "权利要求数量",
            "文献页数",
            "首权字数",
            "IPC",
            "国民经济分类",
            "申请人数量",
            "申请人类型",
            "申请人省市代码",
            "中国申请人地市",
            "中国申请人区县",
            "发明人",
            "引证次数",
            "被引证次数",
            "家族引证次数",
            "家族被引证次数",
            "引证科技文献",
            "简单同族个数",
            "扩展同族个数",
            "inpadoc同族个数",
            "专利寿命（月）",
            "合享价值度",
            "技术稳定性",
            "技术先进性",
            "保护范围"]
need_col_dict = {item: index for index, item in enumerate(need_col)}



def deal_xls_file(file):
    wb = xlrd.open_workbook(file)
    try:
        sht = wb.sheet_by_index(0)
        head = sht.row_values(0)
        key_to_index = get_key_index(head)

        result = []
        cnt = 0
        total_no = 0
        num_rows = sht.nrows
        for cur_row in range(1, num_rows):
            flag = False
            result_data = {}
            data = sht.row_values(cur_row)
            for key in need_col:
                # 保证所有key都出现，因为Key to index可能没有全部的key
                if key in key_to_index:
                    if key == "IPC":
                        ipc_info, green_no = deal_ipc(data[key_to_index[key]])
                        result_data.update(ipc_info)
                        total_no += green_no
                    elif key == "专利类型" and data[key_to_index[key]].strip() == "外观设计":
                        flag = True
                        break
                    else:
                        result_data[key] = data[key_to_index[key]]
                else:
                    result_data[key] = None
            cnt += 1
            MY_DEBUG("cur:{}".format(cnt))
            if flag:
                continue
            result.append(result_data)
        print(total_no)
        LOG("{},green_no:{},total_no:{}".format(file, total_no, sht.nrows-1), "./cnt_log.txt")
        return result, total_no
    finally:
        wb.release_resources()


def deal_xlsx_file(file):
    wb = load_workbook(file, read_only=True)
    try:
        sht = wb.worksheets[0]
        head = sht[1]
        head = [item.value for item in head]
        key_to_index = get_key_index(head)

        result = []
        cnt = 0
        total_no = 0
        for data in sht.iter_rows(min_row=2):
            flag = False
            result_data = {}
            # 保证所有key都在
            for key in need_col:
                # 保证所有key都出现，因为Key to index可能没有全部的key
                if key in key_to_index:
                    if key == "IPC":
                        ipc_info, green_no = deal_ipc(data[key_to_index[key]].value)
                        result_data.update(ipc_info)
                        total_no += green_no
                    elif key == "专利类型" and data[key_to_index[key]].value.strip() == "外观设计":
                        flag = True
                        break
                    else:
                        result_data[key] = data[key_to_index[key]].value
                else:
                    result_data[key] = None
            cnt += 1
            MY_DEBUG("cur:{}".format(cnt))
            if flag:
                continue
            result.append(result_data)
        print(total_no)
        LOG("{},green_no:{},total_no:{}".format(file, total_no, sht.max_row-1), "./cnt_log.txt")
        return result, total_no
    finally:
        wb.close()


def get_key_index(head_str_list):
    num_col = len(head_str_list)

    file_need_col_index = {}
    for i in range(num_col):
        value = str(head_str_list[i]).strip()
        if value in need_col_dict:
            file_need_col_index[value] = i

    return file_need_col_index
